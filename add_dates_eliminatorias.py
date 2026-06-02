import openpyxl
import re

wb = openpyxl.load_workbook('Quiniela_Mundial_2026_Final_vacia.xlsx')
cal_sheet = wb['Calendario de Juegos']
elim_sheet = wb['Eliminatorias']

# 1. Extraer los datos de los Partidos 73 a 104
match_info = {}
current_date = ""

for row in cal_sheet.iter_rows(values_only=True):
    for cell in row:
        if isinstance(cell, str):
            val = cell.strip()
            # If it's a date string
            if "de junio 2026" in val or "de julio 2026" in val:
                current_date = val
            
            # If it's a Knockout match: e.g. "Partido 73 – ..." or "Partido 73 - ..."
            m = re.search(r"Partido\s+(\d{2,3})\s+[–\-]\s+.*?\s+[–\-]\s+(Estadio.*)", val, re.IGNORECASE)
            if m:
                match_num = int(m.group(1))
                location = m.group(2).strip()
                if 73 <= match_num <= 104:
                    match_info[match_num] = f"{current_date} - {location}"

print(f"Extracted {len(match_info)} knockout matches.")
for i in range(73, 105):
    if i not in match_info:
        print(f"Missing match {i}")

# 2. Escribir a la pestaña 'Eliminatorias'
# The row mapping in Eliminatorias for matches 73 to 104
# 73-88: rows 6 to 21
# 89-96: rows 25 to 32
# 97-100: rows 37 to 40
# 101-102: rows 45 to 46
# 103: row 51
# 104: row 56

row_mapping = []
row_mapping.extend(range(6, 22))    # 16 matches
row_mapping.extend(range(25, 33))   # 8 matches
row_mapping.extend(range(37, 41))   # 4 matches
row_mapping.extend(range(45, 47))   # 2 matches
row_mapping.append(51)              # 1 match
row_mapping.append(56)              # 1 match

date_col = 10 # Column J

elim_sheet.cell(row=3, column=date_col, value="Fecha y Estadio")
elim_sheet.column_dimensions[openpyxl.utils.get_column_letter(date_col)].width = 45

for match_num, row_num in zip(range(73, 105), row_mapping):
    info = match_info.get(match_num, "")
    elim_sheet.cell(row=row_num, column=date_col, value=info)

wb.save('Quiniela_Mundial_2026_Final_vacia.xlsx')
print("Successfully added knockout match dates and locations to Eliminatorias.")
