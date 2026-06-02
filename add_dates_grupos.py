import openpyxl
import re

wb = openpyxl.load_workbook('Quiniela_Mundial_2026_Final_vacia.xlsx')
cal_sheet = wb['Calendario de Juegos']

aliases = {
    'Chequia': 'República Checa',
    'Bosnia y Herze.': 'Bosnia y Herzegovina',
    'Curaçao': 'Curazao',
    'Curazao': 'Curazao',
    'Irán': 'RI de Irán',
    'Corea del Sur': 'República de Corea',
    'Qatar': 'Catar',
    'Arabia Saud.': 'Arabia Saudí'
}

# The dictionary should map FROM Calendario name TO Grupos name
# Let's handle special chars in Curaçao
rev_aliases = {v: k for k, v in aliases.items()}

def normalize_name(name):
    name = name.strip()
    return rev_aliases.get(name, name)

matches_info = {}
current_date = ""

for row in cal_sheet.iter_rows(values_only=True):
    for cell in row:
        if isinstance(cell, str):
            val = cell.strip()
            if "de junio" in val or "de julio" in val:
                current_date = val
            elif re.match(r"^\d{2}:\d{2}\s*-", val):
                m = re.search(r"^(\d{2}:\d{2})\s*-\s*(.*?)\s+v\s+(.*?)\s+[–\-]", val)
                if m:
                    time_str = m.group(1)
                    team1 = normalize_name(m.group(2))
                    team2 = normalize_name(m.group(3))
                    matches_info[(team1, team2)] = f"{current_date} - {time_str}"
                    matches_info[(team2, team1)] = f"{current_date} - {time_str}"

grupos_sheet = wb['Grupos']
team1_col = 3
team2_col = 7
date_col = 9

grupos_sheet.cell(row=3, column=date_col, value="Fecha y Hora")
grupos_sheet.column_dimensions[openpyxl.utils.get_column_letter(date_col)].width = 35

found_count = 0
for row in range(4, 100):
    t1 = grupos_sheet.cell(row=row, column=team1_col).value
    t2 = grupos_sheet.cell(row=row, column=team2_col).value
    
    if isinstance(t1, str) and isinstance(t2, str):
        t1 = t1.strip()
        t2 = t2.strip()
        if (t1, t2) in matches_info:
            grupos_sheet.cell(row=row, column=date_col, value=matches_info[(t1, t2)])
            # Match styling from a nearby cell if needed, but not strictly necessary
            found_count += 1
        elif t1 and t2 and t1 != "Equipo 1" and t2 != "Equipo 2":
            print(f"Warning: match not found ({t1} vs {t2})")

wb.save('Quiniela_Mundial_2026_Final_vacia.xlsx')
print(f"Dates added successfully. Mapped {found_count} matches.")
