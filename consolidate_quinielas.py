#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
SISTEMA DE CONSOLIDACIÓN Y EVALUACIÓN DE QUINIELAS - MUNDIAL DE FÚTBOL 2026
--------------------------------------------------------------------------
Este script permite automatizar la evaluación de las quinielas de fútbol de los
participantes, comparando sus predicciones contra un archivo Maestro (controlado
por el administrador con los resultados reales) y generando una tabla de 
clasificación (Leaderboard) espectacularmente diseñada en Excel y un resumen de texto.

Creado por Antigravity (Google DeepMind).
"""

import os
import sys
import unicodedata
import argparse
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Reconfigurar stdout y stderr para forzar UTF-8 y evitar errores de encoding en Windows CMD
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='backslashreplace')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8', errors='backslashreplace')

# ==============================================================================
# CONFIGURACIÓN DEL SISTEMA DE PUNTUACIÓN (AJUSTABLE POR EL ADMINISTRADOR)
# ==============================================================================
POINTS_CONFIG = {
    # Fase de Grupos
    "GROUP_EXACT": 3,      # Puntos por marcador exacto (ej. predice 2-1, resultado 2-1)
    "GROUP_OUTCOME": 1,    # Puntos por acertar ganador/empate pero no marcador exacto

    # Marcadores en Eliminatorias (Solo si los equipos del partido coinciden)
    "KO_MATCH_EXACT": 5,   # Puntos por marcador exacto en eliminatorias
    "KO_MATCH_OUTCOME": 2, # Puntos por acertar ganador en eliminatorias

    # Puntos por equipos que clasifican a cada ronda (Sistema Bracket de March Madness)
    "TEAM_R32": 2,         # Por cada selección que logra entrar a Dieciseisavos (32 equipos)
    "TEAM_R16": 4,         # Por cada selección que logra entrar a Octavos (16 equipos)
    "TEAM_QF": 6,          # Por cada selección que logra entrar a Cuartos de Final (8 equipos)
    "TEAM_SF": 8,          # Por cada selección que logra entrar a Semifinales (4 equipos)
    "TEAM_T3P": 10,        # Por cada selección que logra entrar al partido de 3er Lugar (2 equipos)
    "TEAM_FINAL": 12,      # Por cada selección que logra entrar a la Gran Final (2 equipos)

    # Puntos por aciertos en el Podio de Honor
    "PODIUM_CHAMPION": 15, # Por acertar al Campeón Mundial exacto 🏆
    "PODIUM_RUNNERUP": 10, # Por acertar al Subcampeón exacto 🥈
    "PODIUM_THIRD": 10,    # Por acertar al Tercer Lugar exacto 🥉
}

# ==============================================================================
# FUNCIONES AUXILIARES DE NORMALIZACIÓN Y LECTURA
# ==============================================================================
def normalize_team_name(name):
    """
    Normaliza los nombres de los equipos para evitar errores de comparación
    debido a acentos, mayúsculas, espacios extra o caracteres especiales.
    """
    if name is None:
        return ""
    # Convertir a cadena, pasar a minúsculas y quitar espacios en extremos
    s = str(name).strip().lower()
    # Eliminar acentos y diacríticos (ej. 'México' -> 'mexico')
    s = "".join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    # Reemplazar puntos, guiones y abreviaciones comunes
    s = s.replace(".", "").replace("-", " ").replace("  ", " ")
    return s

def clean_score(val):
    """Convierte un valor de celda a un entero de forma segura, si es posible."""
    if val is None or val == "" or val == "-":
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None

# ==============================================================================
# PROCESAMIENTO DE LAS PESTAÑAS DEL EXCEL
# ==============================================================================
def process_group_stage(ws, is_master=False):
    """
    Extrae todos los marcadores de la pestaña 'Grupos'.
    Retorna un diccionario de partidos {(Equipo1, Equipo2): (Goles1, Goles2)}
    y una lista secuencial de partidos.
    """
    matches = {}
    match_list = []
    
    # Hay 12 grupos (A-L), cada uno espaciado exactamente 8 filas
    # Grupo A empieza en fila 6, Grupo B en 14, Grupo C en 22, ..., Grupo L en 94.
    for g in range(12):
        start_row = 6 + (8 * g)
        for m in range(6):
            r = start_row + m
            t1 = ws.cell(row=r, column=3).value  # Columna C: Equipo 1
            s1 = clean_score(ws.cell(row=r, column=4).value)  # Columna D: Goles 1
            s2 = clean_score(ws.cell(row=r, column=6).value)  # Columna F: Goles 2
            t2 = ws.cell(row=r, column=7).value  # Columna G: Equipo 2
            
            if t1 and t2:
                norm_t1 = normalize_team_name(t1)
                norm_t2 = normalize_team_name(t2)
                
                match_data = {
                    "row": r,
                    "team1": t1,
                    "team2": t2,
                    "norm_team1": norm_t1,
                    "norm_team2": norm_t2,
                    "score1": s1,
                    "score2": s2
                }
                match_list.append(match_data)
                
                # Guardar en diccionario indexado por equipos normalizados
                matches[(norm_t1, norm_t2)] = match_data
                
    return matches, match_list

def process_knockout_stage(ws):
    """
    Extrae los datos de la pestaña 'Eliminatorias'.
    Retorna un diccionario con los equipos avanzados por ronda y los resultados de partidos.
    """
    # 1. Extraer los equipos avanzados de cada fase (según la estructura real de celdas)
    advances = {
        "r32": [],    # Dieciseisavos (32 equipos en filas 6-21)
        "r16": [],    # Octavos (16 equipos en filas 25-32)
        "qf": [],     # Cuartos (8 equipos en filas 37-40)
        "sf": [],     # Semifinales (4 equipos en filas 45-46)
        "t3p": [],    # Disputan 3er Lugar (2 equipos en fila 51)
        "final": [],  # Finalistas (2 equipos en fila 56)
        "podium": {}  # Campeón, Subcampeón, 3er Lugar
    }

    def get_raw_val(r, c):
        val = ws.cell(row=r, column=c).value
        return str(val).strip() if val else ""

    # Dieciseisavos (filas 6 a 21)
    for r in range(6, 22):
        t1, t2 = get_raw_val(r, 2), get_raw_val(r, 6) # Col B y Col F
        if t1: advances["r32"].append(t1)
        if t2: advances["r32"].append(t2)

    # Octavos (filas 25 a 32)
    for r in range(25, 33):
        t1, t2 = get_raw_val(r, 2), get_raw_val(r, 6)
        if t1: advances["r16"].append(t1)
        if t2: advances["r16"].append(t2)

    # Cuartos (filas 37 a 40)
    for r in range(37, 41):
        t1, t2 = get_raw_val(r, 2), get_raw_val(r, 6)
        if t1: advances["qf"].append(t1)
        if t2: advances["qf"].append(t2)

    # Semifinales (filas 45 a 46)
    for r in range(45, 47):
        t1, t2 = get_raw_val(r, 2), get_raw_val(r, 6)
        if t1: advances["sf"].append(t1)
        if t2: advances["sf"].append(t2)

    # Disputan Tercer Lugar (fila 51)
    t1_t3, t2_t3 = get_raw_val(51, 2), get_raw_val(51, 6)
    if t1_t3: advances["t3p"].append(t1_t3)
    if t2_t3: advances["t3p"].append(t2_t3)

    # Gran Final (fila 56)
    t1_f, t2_f = get_raw_val(56, 2), get_raw_val(56, 6)
    if t1_f: advances["final"].append(t1_f)
    if t2_f: advances["final"].append(t2_f)

    # Podio de Honor
    advances["podium"]["champion"] = get_raw_val(60, 6)   # F60
    advances["podium"]["runnerup"] = get_raw_val(61, 6)   # F61
    advances["podium"]["third"] = get_raw_val(62, 6)      # F62

    # 2. Extraer los marcadores de los partidos individuales de Eliminatorias
    ko_matches = {}
    
    # Lista de tuplas con (rango_filas, fase_nombre)
    ko_phases = [
        (range(6, 22), "r32"),
        (range(25, 33), "r16"),
        (range(37, 41), "qf"),
        (range(45, 47), "sf"),
        ([51], "t3p"),
        ([56], "final")
    ]
    
    for row_range, phase_name in ko_phases:
        for r in row_range:
            t1 = ws.cell(row=r, column=2).value  # Columna B: Equipo A
            s1 = clean_score(ws.cell(row=r, column=3).value)  # Columna C: Goles A
            s2 = clean_score(ws.cell(row=r, column=5).value)  # Columna E: Goles B
            t2 = ws.cell(row=r, column=6).value  # Columna F: Equipo B
            winner = ws.cell(row=r, column=7).value  # Columna G: Clasifica
            
            if t1 and t2:
                norm_t1 = normalize_team_name(t1)
                norm_t2 = normalize_team_name(t2)
                ko_matches[(norm_t1, norm_t2)] = {
                    "row": r,
                    "team1": t1,
                    "team2": t2,
                    "score1": s1,
                    "score2": s2,
                    "winner": winner,
                    "phase": phase_name
                }
                
    return advances, ko_matches

# ==============================================================================
# ALGORITMO PRINCIPAL DE EVALUACIÓN DE PUNTOS
# ==============================================================================
def evaluate_participant(participant_path, master_groups, master_groups_list, master_ko_advances, master_ko_matches):
    """
    Compara las predicciones de un participante contra el Maestro
    y calcula el desglose detallado de puntos.
    """
    # Cargar el archivo con data_only=True para evaluar los valores calculados por fórmulas
    wb = openpyxl.load_workbook(participant_path, data_only=True)
    
    # 1. Evaluar Fase de Grupos
    group_pts = 0
    group_exact_matches = 0
    group_outcome_matches = 0
    group_matches_played = 0
    
    part_groups, _ = process_group_stage(wb['Grupos'])
    
    for master_key, master_match in master_groups.items():
        m_s1, m_s2 = master_match["score1"], master_match["score2"]
        
        # Solo evaluar si el administrador ya cargó un resultado real en el Maestro
        if m_s1 is not None and m_s2 is not None:
            group_matches_played += 1
            
            # Buscar el partido en el archivo del participante (por equipos normalizados)
            p_match = part_groups.get(master_key)
            if p_match:
                p_s1, p_s2 = p_match["score1"], p_match["score2"]
                
                # Si el participante no llenó el partido, se salta con 0 puntos
                if p_s1 is not None and p_s2 is not None:
                    # Caso 1: Acierto Exacto de Marcador
                    if p_s1 == m_s1 and p_s2 == m_s2:
                        group_pts += POINTS_CONFIG["GROUP_EXACT"]
                        group_exact_matches += 1
                    # Caso 2: Acierto de Ganador/Empate
                    elif (p_s1 > p_s2 and m_s1 > m_s2) or (p_s1 < p_s2 and m_s1 < m_s2) or (p_s1 == p_s2 and m_s1 == m_s2):
                        group_pts += POINTS_CONFIG["GROUP_OUTCOME"]
                        group_outcome_matches += 1
    
    # 2. Evaluar Avanzados de Eliminatorias (March Madness Bracket System)
    part_ko_advances, part_ko_matches = process_knockout_stage(wb['Eliminatorias'])
    
    r32_pts = 0
    r16_pts = 0
    qf_pts = 0
    sf_pts = 0
    t3p_pts = 0
    final_pts = 0
    podium_pts = 0
    
    # Normalizar conjuntos del participante para una comparación robusta
    def get_norm_set(team_list):
        return {normalize_team_name(t) for t in team_list if t and str(t).strip() != ""}
        
    part_r32_set = get_norm_set(part_ko_advances["r32"])
    part_r16_set = get_norm_set(part_ko_advances["r16"])
    part_qf_set = get_norm_set(part_ko_advances["qf"])
    part_sf_set = get_norm_set(part_ko_advances["sf"])
    part_t3p_set = get_norm_set(part_ko_advances["t3p"])
    part_final_set = get_norm_set(part_ko_advances["final"])
    
    # Evaluar R32 (Dieciseisavos)
    for team in master_ko_advances["r32"]:
        norm_t = normalize_team_name(team)
        if norm_t and norm_t in part_r32_set:
            r32_pts += POINTS_CONFIG["TEAM_R32"]
            
    # Evaluar R16 (Octavos)
    for team in master_ko_advances["r16"]:
        norm_t = normalize_team_name(team)
        if norm_t and norm_t in part_r16_set:
            r16_pts += POINTS_CONFIG["TEAM_R16"]
            
    # Evaluar QF (Cuartos)
    for team in master_ko_advances["qf"]:
        norm_t = normalize_team_name(team)
        if norm_t and norm_t in part_qf_set:
            qf_pts += POINTS_CONFIG["TEAM_QF"]
            
    # Evaluar SF (Semifinales)
    for team in master_ko_advances["sf"]:
        norm_t = normalize_team_name(team)
        if norm_t and norm_t in part_sf_set:
            sf_pts += POINTS_CONFIG["TEAM_SF"]

    # Evaluar disputantes de 3er Lugar
    for team in master_ko_advances["t3p"]:
        norm_t = normalize_team_name(team)
        if norm_t and norm_t in part_t3p_set:
            t3p_pts += POINTS_CONFIG["TEAM_T3P"]
            
    # Evaluar Finalistas
    for team in master_ko_advances["final"]:
        norm_t = normalize_team_name(team)
        if norm_t and norm_t in part_final_set:
            final_pts += POINTS_CONFIG["TEAM_FINAL"]
            
    # Evaluar Podio de Honor
    # Campeón
    m_camp = normalize_team_name(master_ko_advances["podium"]["champion"])
    p_camp = normalize_team_name(part_ko_advances["podium"]["champion"])
    if m_camp and m_camp == p_camp:
        podium_pts += POINTS_CONFIG["PODIUM_CHAMPION"]
        
    # Subcampeón
    m_sub = normalize_team_name(master_ko_advances["podium"]["runnerup"])
    p_sub = normalize_team_name(part_ko_advances["podium"]["runnerup"])
    if m_sub and m_sub == p_sub:
        podium_pts += POINTS_CONFIG["PODIUM_RUNNERUP"]
        
    # Tercer Lugar
    m_third = normalize_team_name(master_ko_advances["podium"]["third"])
    p_third = normalize_team_name(part_ko_advances["podium"]["third"])
    if m_third and m_third == p_third:
        podium_pts += POINTS_CONFIG["PODIUM_THIRD"]

    # 3. Evaluar marcadores de Eliminatorias (si coinciden los partidos disputados)
    ko_match_score_pts = 0
    ko_exact_matches = 0
    ko_outcome_matches = 0
    
    for master_key, m_match in master_ko_matches.items():
        m_s1, m_s2 = m_match["score1"], m_match["score2"]
        if m_s1 is not None and m_s2 is not None:
            # Buscar en el participante
            p_match = part_ko_matches.get(master_key)
            if not p_match:
                p_match = part_ko_matches.get((master_key[1], master_key[0]))
                if p_match:
                    p_s1, p_s2 = p_match["score2"], p_match["score1"]
                else:
                    p_s1, p_s2 = None, None
            else:
                p_s1, p_s2 = p_match["score1"], p_match["score2"]
                
            if p_s1 is not None and p_s2 is not None:
                if p_s1 == m_s1 and p_s2 == m_s2:
                    ko_match_score_pts += POINTS_CONFIG["KO_MATCH_EXACT"]
                    ko_exact_matches += 1
                elif (p_s1 > p_s2 and m_s1 > m_s2) or (p_s1 < p_s2 and m_s1 < m_s2) or (p_s1 == p_s2 and m_s1 == m_s2):
                    ko_match_score_pts += POINTS_CONFIG["KO_MATCH_OUTCOME"]
                    ko_outcome_matches += 1

    # Totalizar puntos
    total_pts = (group_pts + r32_pts + r16_pts + qf_pts + sf_pts + 
                 t3p_pts + final_pts + podium_pts + ko_match_score_pts)
    
    exact_aciertos_totales = group_exact_matches + ko_exact_matches
    
    wb.close()
    
    return {
        "total_points": total_pts,
        "group_pts": group_pts,
        "r32_pts": r32_pts,
        "r16_pts": r16_pts,
        "qf_pts": qf_pts,
        "sf_pts": sf_pts,
        "t3p_pts": t3p_pts,
        "final_pts": final_pts,
        "podium_pts": podium_pts,
        "ko_match_score_pts": ko_match_score_pts,
        "group_exact_matches": group_exact_matches,
        "group_outcome_matches": group_outcome_matches,
        "ko_exact_matches": ko_exact_matches,
        "ko_outcome_matches": ko_outcome_matches,
        "exact_aciertos_totales": exact_aciertos_totales,
        "group_matches_played": group_matches_played
    }

# ==============================================================================
# CREACIÓN DE LA PLANTILLA DEL LEADERBOARD CON DISEÑO PREMIUM
# ==============================================================================
def create_premium_leaderboard(results, output_path, max_possible_pts, master_groups_list=None, master_ko_matches=None):
    """
    Genera un archivo de Excel espectacular, usando una paleta de colores
    coherente (Verdes/Esmeraldas) que coincide con el diseño de la quiniela.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clasificación General"
    
    # Asegurar que se muestren las líneas de cuadrícula
    ws.views.sheetView[0].showGridLines = True
    
    # 1. Definición del Sistema de Diseño (Estilos)
    font_family = "Segoe UI"
    
    # Fills (Rellenos)
    fill_header_dark = PatternFill(fill_type="solid", start_color="0F4C43", end_color="0F4C43") # Verde Oscuro Premium
    fill_subheader = PatternFill(fill_type="solid", start_color="E3ECEB", end_color="E3ECEB")    # Verde Menta Claro
    fill_zebra_light = PatternFill(fill_type="solid", start_color="F7FAF9", end_color="F7FAF9")  # Cebra ultra suave
    fill_gold = PatternFill(fill_type="solid", start_color="FFF3CD", end_color="FFF3CD")         # Oro/Amarillo suave
    fill_silver = PatternFill(fill_type="solid", start_color="E2E3E5", end_color="E2E3E5")       # Plata suave
    fill_bronze = PatternFill(fill_type="solid", start_color="F8D7DA", end_color="F8D7DA")       # Bronce/Rosa suave
    
    # Fonts (Fuentes)
    font_title = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    font_subtitle = Font(name=font_family, size=10, italic=True, color="0F4C43")
    font_header = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    font_data_bold = Font(name=font_family, size=10, bold=True, color="0F4C43")
    font_data_normal = Font(name=font_family, size=10, color="333333")
    font_top3 = Font(name=font_family, size=10, bold=True, color="0F4C43")
    
    # Borders (Bordes)
    thin_side = Side(border_style="thin", color="CCCCCC")
    thick_bottom = Side(border_style="medium", color="0F4C43")
    border_cell = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    border_header = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thick_bottom)
    
    # Alignments (Alineaciones)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    
    # 2. Construcción del Banner de Título
    ws.merge_cells("A1:L2")
    title_cell = ws["A1"]
    title_cell.value = "CLASIFICACIÓN DE PARTICIPANTES - QUINIELA MUNDIAL 2026"
    title_cell.font = font_title
    title_cell.fill = fill_header_dark
    title_cell.alignment = align_center
    
    # Ajustar alto de las filas de título
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20
    
    # Fila de Subtítulo (Metadatos)
    ws.merge_cells("A3:L3")
    sub_cell = ws["A3"]
    sub_cell.value = f"Actualizado el: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Puntos Máximos Posibles Disputados: {max_possible_pts} | Sistema de puntuación automatizado"
    sub_cell.font = font_subtitle
    sub_cell.fill = fill_subheader
    sub_cell.alignment = align_center
    ws.row_dimensions[3].height = 22
    
    ws.row_dimensions[4].height = 15  # Separador en blanco
    
    # 3. Cabecera de la Tabla
    headers = [
        "Lugar", "Nombre del Participante", "Puntos Totales", "% Efectividad", 
        "Aciertos Exactos", "Puntos Grupos", "Puntos R32", "Puntos Octavos", 
        "Puntos Cuartos", "Puntos Semis", "Puntos Final", "Puntos Podio"
    ]
    
    ws.row_dimensions[5].height = 28
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = text
        cell.font = font_header
        cell.fill = fill_header_dark
        cell.alignment = align_center
        cell.border = border_header
        
    # 4. Inserción de Datos y Formato Dinámico
    # Nota: results viene ordenado por puntos totales (descendente) y luego aciertos exactos (descendente)
    for idx, r in enumerate(results, 1):
        row_num = 5 + idx
        ws.row_dimensions[row_num].height = 22
        
        # Asignar estilo de fila especial para el Top 3 (Oro, Plata, Bronce)
        if idx == 1:
            row_fill = fill_gold
            medal = "🏆 "
            row_font = font_top3
        elif idx == 2:
            row_fill = fill_silver
            medal = "🥈 "
            row_font = font_top3
        elif idx == 3:
            row_fill = fill_bronze
            medal = "🥉 "
            row_font = font_top3
        else:
            row_fill = fill_zebra_light if idx % 2 == 0 else PatternFill(fill_type=None)
            medal = ""
            row_font = font_data_normal
            
        # Rellenar cada columna
        # Col 1: Lugar
        c = ws.cell(row=row_num, column=1, value=idx)
        c.font = Font(name=font_family, size=10, bold=True)
        c.alignment = align_center
        c.border = border_cell
        if row_fill.fill_type: c.fill = row_fill
        
        # Col 2: Nombre del Participante
        c = ws.cell(row=row_num, column=2, value=f"{medal}{r['name']}")
        c.font = row_font
        c.alignment = align_left
        c.border = border_cell
        if row_fill.fill_type: c.fill = row_fill
        
        # Col 3: Puntos Totales (Destacado en Negrita siempre)
        c = ws.cell(row=row_num, column=3, value=r["total_points"])
        c.font = font_data_bold
        c.alignment = align_center
        c.border = border_cell
        if row_fill.fill_type: c.fill = row_fill
        
        # Col 4: Efectividad
        # Calcular porcentaje respecto al máximo teórico real jugado hasta ahora
        effectiveness = (r["total_points"] / max_possible_pts) if max_possible_pts > 0 else 0.0
        c = ws.cell(row=row_num, column=4, value=effectiveness)
        c.font = row_font
        c.alignment = align_center
        c.number_format = '0.0%'
        c.border = border_cell
        if row_fill.fill_type: c.fill = row_fill
        
        # Col 5: Aciertos Exactos
        c = ws.cell(row=row_num, column=5, value=r["exact_aciertos_totales"])
        c.font = row_font
        c.alignment = align_center
        c.border = border_cell
        if row_fill.fill_type: c.fill = row_fill
        
        # Desglose de puntos por ronda
        breakdown_keys = [
            "group_pts", "r32_pts", "r16_pts", "qf_pts", "sf_pts", "final_pts", "podium_pts"
        ]
        for offset, key in enumerate(breakdown_keys, 6):
            c = ws.cell(row=row_num, column=offset, value=r[key])
            c.font = font_data_normal
            c.alignment = align_center
            c.border = border_cell
            if row_fill.fill_type: c.fill = row_fill

    # 5. Ajuste Automático de Ancho de Columnas
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        # No calcular el ancho con las filas combinadas del Banner de Título (1 a 3)
        for cell in col[4:]:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        
        # Ajustar con un margen de seguridad
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    # Columna del nombre necesita un ancho especial más generoso
    ws.column_dimensions["B"].width = 30
    
    # ==============================================================================
    # PESTAÑA DE RESULTADOS OFICIALES
    # ==============================================================================
    if master_groups_list is not None and master_ko_matches is not None:
        ws_res = wb.create_sheet(title="Resultados Oficiales")
        ws_res.views.sheetView[0].showGridLines = False
        
        # Titulo
        ws_res.merge_cells("A1:C2")
        title_res = ws_res["A1"]
        title_res.value = "RESULTADOS OFICIALES DEL MUNDIAL"
        title_res.font = font_title
        title_res.fill = fill_header_dark
        title_res.alignment = align_center
        ws_res.row_dimensions[1].height = 20
        ws_res.row_dimensions[2].height = 20
        
        ws_res.row_dimensions[3].height = 15
        
        # Cabeceras
        for col_idx, text in enumerate(["Equipo 1", "Marcador", "Equipo 2"], 1):
            c = ws_res.cell(row=4, column=col_idx, value=text)
            c.font = font_header
            c.fill = fill_header_dark
            c.alignment = align_center
            c.border = border_header
            
        current_row = 5
        
        # Filtros: solo partidos jugados
        played_groups = [m for m in master_groups_list if m["score1"] is not None and m["score2"] is not None]
        played_ko = [m for m in master_ko_matches.values() if m["score1"] is not None and m["score2"] is not None]
        
        def write_section_header(title, row):
            ws_res.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            c = ws_res.cell(row=row, column=1, value=title)
            c.font = font_data_bold
            c.fill = fill_subheader
            c.alignment = align_center
            c.border = border_cell
        
        def write_match(m, row):
            c1 = ws_res.cell(row=row, column=1, value=m['team1'])
            c2 = ws_res.cell(row=row, column=2, value=f"{m['score1']} - {m['score2']}")
            c3 = ws_res.cell(row=row, column=3, value=m['team2'])
            
            for c in [c1, c2, c3]:
                c.font = font_data_normal
                c.alignment = align_center
                c.border = border_cell
                if row % 2 == 0:
                    c.fill = fill_zebra_light

        # Escribir Grupos
        if played_groups:
            write_section_header("Fase de Grupos", current_row)
            current_row += 1
            for m in played_groups:
                write_match(m, current_row)
                current_row += 1
                
        # Espacio
        current_row += 1
        
        # Escribir Eliminatorias
        if played_ko:
            write_section_header("Fase Eliminatoria", current_row)
            current_row += 1
            for m in played_ko:
                write_match(m, current_row)
                current_row += 1
                
        # Ancho de columnas
        ws_res.column_dimensions["A"].width = 25
        ws_res.column_dimensions["B"].width = 15
        ws_res.column_dimensions["C"].width = 25
    
    wb.save(output_path)
    wb.close()

# ==============================================================================
# GENERACIÓN DE REPORTE PARA WHATSAPP
# ==============================================================================
def generate_whatsapp_report(results, max_possible_pts):
    """
    Genera un texto perfectamente formateado listo para copiar y pegar en WhatsApp
    para mantener informados a todos los miembros de la quiniela.
    """
    now_str = datetime.now().strftime('%d/%m/%Y a las %H:%M')
    report = []
    report.append(f"🏆 *CLASIFICACIÓN QUINIELA MUNDIAL 2026* 🏆")
    report.append(f"📊 _Actualizado: {now_str}_")
    report.append(f"🎯 _Puntos Máximos Disputados hasta hoy: {max_possible_pts}_")
    report.append(f"-------------------------------------------")
    
    for idx, r in enumerate(results, 1):
        effectiveness_pct = (r["total_points"] / max_possible_pts * 100) if max_possible_pts > 0 else 0.0
        
        if idx == 1:
            emoji = "🥇"
            prefix = f"*{idx}. {emoji} {r['name'].upper()}*"
        elif idx == 2:
            emoji = "🥈"
            prefix = f"*{idx}. {emoji} {r['name']}*"
        elif idx == 3:
            emoji = "🥉"
            prefix = f"*{idx}. {emoji} {r['name']}*"
        else:
            emoji = "🏃‍♂️"
            prefix = f"{idx}. {r['name']}"
            
        points_str = f"*{r['total_points']} pts*"
        details = f"({r['exact_aciertos_totales']} marcadores exactos | {effectiveness_pct:.1f}% efectividad)"
        report.append(f"{prefix} ➔ {points_str} {details}")
        
    report.append(f"-------------------------------------------")
    report.append(f"🔥 ¡Felicidades al líder temporal! ¿Quién ganará? 🍿🚀")
    
    return "\n".join(report)

# ==============================================================================
# PROCESAMIENTO DEL ARCHIVO MAESTRO
# ==============================================================================
def process_master_file(master_file):
    """
    Carga el archivo Maestro, extrae los resultados oficiales y calcula 
    los puntos máximos disputados hasta el momento.
    """
    wb_master = openpyxl.load_workbook(master_file, data_only=True)
    master_groups, master_groups_list = process_group_stage(wb_master['Grupos'], is_master=True)
    master_ko_advances, master_ko_matches = process_knockout_stage(wb_master['Eliminatorias'])
    wb_master.close()
    
    max_possible_pts = 0
    
    # Fase de Grupos
    group_matches_played = sum(1 for m in master_groups_list if m["score1"] is not None and m["score2"] is not None)
    max_possible_pts += group_matches_played * POINTS_CONFIG["GROUP_EXACT"]
    
    # Eliminatorias - Marcadores de partidos disputados
    ko_matches_played = sum(1 for m in master_ko_matches.values() if m["score1"] is not None and m["score2"] is not None)
    max_possible_pts += ko_matches_played * POINTS_CONFIG["KO_MATCH_EXACT"]
    
    # Eliminatorias - Equipos avanzados
    r32_played = sum(1 for t in master_ko_advances["r32"] if t and str(t).strip() != "")
    max_possible_pts += r32_played * POINTS_CONFIG["TEAM_R32"]
    
    r16_played = sum(1 for t in master_ko_advances["r16"] if t and str(t).strip() != "")
    max_possible_pts += r16_played * POINTS_CONFIG["TEAM_R16"]
    
    qf_played = sum(1 for t in master_ko_advances["qf"] if t and str(t).strip() != "")
    max_possible_pts += qf_played * POINTS_CONFIG["TEAM_QF"]
    
    sf_played = sum(1 for t in master_ko_advances["sf"] if t and str(t).strip() != "")
    max_possible_pts += sf_played * POINTS_CONFIG["TEAM_SF"]
    
    t3p_played = sum(1 for t in master_ko_advances["t3p"] if t and str(t).strip() != "")
    max_possible_pts += t3p_played * POINTS_CONFIG["TEAM_T3P"]
    
    final_played = sum(1 for t in master_ko_advances["final"] if t and str(t).strip() != "")
    max_possible_pts += final_played * POINTS_CONFIG["TEAM_FINAL"]
    
    # Podio
    if master_ko_advances["podium"]["champion"]:
        max_possible_pts += POINTS_CONFIG["PODIUM_CHAMPION"]
    if master_ko_advances["podium"]["runnerup"]:
        max_possible_pts += POINTS_CONFIG["PODIUM_RUNNERUP"]
    if master_ko_advances["podium"]["third"]:
        max_possible_pts += POINTS_CONFIG["PODIUM_THIRD"]
        
    return master_groups, master_groups_list, master_ko_advances, master_ko_matches, max_possible_pts, group_matches_played

# ==============================================================================
# ORQUESTACIÓN GENERAL Y ENTRADA DE LÍNEA DE COMANDOS
# ==============================================================================
def main():
    parser = argparse.ArgumentParser(
        description="Evalúa y consolida las quinielas del Mundial 2026."
    )
    parser.add_argument(
        "--master", 
        default=r"C:\Users\salvador.velasco\Downloads\Quiniela_Mundial_2026_Totalmente_Corregida.xlsx",
        help="Ruta absoluta del archivo Master con resultados reales."
    )
    parser.add_argument(
        "--folder", 
        default=r"C:\Users\salvador.velasco\Downloads\Participantes",
        help="Ruta de la carpeta que contiene las quinielas de los participantes."
    )
    parser.add_argument(
        "--output", 
        default=r"C:\Users\salvador.velasco\Downloads\Clasificacion_Quiniela_Mundial_2026.xlsx",
        help="Ruta del archivo Excel de salida con la tabla de clasificación."
    )
    
    args = parser.parse_args()
    
    print("\n======================================================================")
    print(" INICIANDO CONSOLIDACION Y EVALUACION DE QUINIELAS MUNDIAL 2026")
    print("======================================================================\n")
    
    # 1. Validar la existencia del archivo Maestro
    if not os.path.exists(args.master):
        print(f"ERROR: El archivo Maestro no existe en la ruta:\n   {args.master}")
        sys.exit(1)
        
    # 2. Cargar datos del Maestro
    print(f"Cargando archivo Maestro con resultados oficiales...")
    try:
        master_groups, master_groups_list, master_ko_advances, master_ko_matches, max_possible_pts, group_matches_played = process_master_file(args.master)
    except Exception as e:
        print(f"ERROR: No se pudo abrir o procesar el archivo Maestro: {e}")
        sys.exit(1)
        
    print(f"Resultados oficiales cargados. Partidos grupales oficiales jugados: {group_matches_played}.")
    print(f"Puntos maximos disputados hasta el momento: {max_possible_pts} pts.")
    
    # 3. Validar y procesar carpeta de participantes
    if not os.path.exists(args.folder):
        print(f"Creando carpeta para participantes en: {args.folder}")
        os.makedirs(args.folder)
        print("Pon en esa carpeta los archivos Excel llenados por los participantes.")
        print("Asegúrate de que sus archivos tengan nombres descriptivos (ej: 'Quiniela_Sofia.xlsx')")
        print("Vuelve a ejecutar el script una vez que coloques los archivos de los participantes.")
        sys.exit(0)
        
    participant_files = [
        f for f in os.listdir(args.folder) 
        if f.endswith('.xlsx') and not f.startswith('~$') and not f.lower().count('master') and not f.lower().count('clasificacion')
    ]
    
    if not participant_files:
        print(f"AVISO: No se encontraron archivos de participantes en la carpeta:\n   {args.folder}")
        print("Por favor, coloca los archivos Excel de tus amigos o participantes allí y vuelve a ejecutar.")
        sys.exit(0)
        
    print(f"Se encontraron {len(participant_files)} quinielas de participantes listos para evaluar.")
    
    # 4. Procesar uno a uno los archivos de participantes
    participant_results = []
    
    for filename in participant_files:
        path = os.path.join(args.folder, filename)
        
        # Extraer nombre del participante desde el nombre del archivo
        p_name = filename.replace('.xlsx', '').replace('Quiniela_Mundial_2026_', '').replace('Quiniela_Mundial_', '').replace('Quiniela_', '')
        p_name = p_name.replace('_', ' ').strip()
        
        print(f"Evaluando quiniela de: {p_name} ({filename})...")
        
        try:
            scores = evaluate_participant(path, master_groups, master_groups_list, master_ko_advances, master_ko_matches)
            scores["name"] = p_name
            scores["file"] = filename
            participant_results.append(scores)
        except Exception as e:
            print(f"ERROR: No se pudo evaluar el archivo de {p_name}: {e}")
            import traceback
            traceback.print_exc()
            
    if not participant_results:
        print("ERROR: No se pudieron evaluar de forma exitosa los archivos de los participantes.")
        sys.exit(1)
        
    # 5. Ordenar la tabla de clasificación
    participant_results.sort(
        key=lambda x: (-x["total_points"], -x["exact_aciertos_totales"], x["name"].lower())
    )
    
    # 6. Generar el archivo de clasificación espectacular de Excel
    print(f"\nGenerando el archivo Excel Leaderboard de Clasificación General...")
    try:
        create_premium_leaderboard(participant_results, args.output, max_possible_pts)
        print(f"¡Archivo de clasificación de Excel creado con éxito en:\n   {args.output}")
    except Exception as e:
        print(f"ERROR: No se pudo generar la clasificación de Excel: {e}")
        
    # 7. Generar reporte para WhatsApp
    print(f"\nGenerando reporte listo para WhatsApp...")
    report_text = generate_whatsapp_report(participant_results, max_possible_pts)
    
    # Guardar reporte de texto en un archivo
    report_path = os.path.join(os.path.dirname(args.output), "Reporte_WhatsApp_Quiniela_2026.txt")
    with open(report_path, "w", encoding="utf-8") as f:
        f.write(report_text)
        
    print(f"¡Reporte de texto guardado en:\n   {report_path}")
    print("\n" + "="*50)
    print(" COPIA Y PEGA ESTO EN TU GRUPO DE WHATSAPP:")
    print("="*50)
    print(report_text)
    print("="*50 + "\n")
    
if __name__ == "__main__":
    main()
