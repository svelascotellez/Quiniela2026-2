import openpyxl

source_wb = openpyxl.load_workbook('ranking_fifa_masculino.xlsx')
source_sheet = source_wb.active

target_wb = openpyxl.load_workbook('Quiniela_Mundial_2026_Final_vacia.xlsx')
if "Ranking FIFA" in target_wb.sheetnames:
    del target_wb["Ranking FIFA"]
target_sheet = target_wb.create_sheet(title="Ranking FIFA")

for row in source_sheet:
    for cell in row:
        new_cell = target_sheet[cell.coordinate]
        new_cell.value = cell.value
        
        if cell.has_style:
            new_cell.font = cell.font.copy()
            new_cell.border = cell.border.copy()
            new_cell.fill = cell.fill.copy()
            new_cell.number_format = cell.number_format
            new_cell.protection = cell.protection.copy()
            new_cell.alignment = cell.alignment.copy()

target_wb.save('Quiniela_Mundial_2026_Final_vacia.xlsx')
print("Ranking FIFA added to template successfully.")
